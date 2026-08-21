"""Run a resumable, cache-free reliability benchmark against real URLs.

Raw results are intentionally written below ``.local`` by the caller: they can
contain public comment text and transient URL tokens and must not be committed.
"""

from __future__ import annotations

import argparse
import asyncio
from collections import Counter
from contextlib import asynccontextmanager
from contextvars import ContextVar
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from hashlib import sha256
from io import BytesIO
import json
from pathlib import Path
from time import monotonic
from typing import Any, AsyncIterator, Literal

from openpyxl import load_workbook

from app.core.config import Settings
from app.core.logging import configure_logging
from app.crawlers.browser_fallback import BrowserFallback, BrowserFallbackSettings
from app.crawlers.engagement import EngagementCrawler
from app.crawlers.http_client import CurlAsyncHttpClient
from app.crawlers.platform_session import PlatformSessionStore
from app.crawlers.platforms.registry import identify_url
from app.crawlers.proxy_provider import AsyncDailiProxyPool, AsyncRequestRateLimiter
from app.models.engagement import CommentPageResult, InteractionResult


Operation = Literal["interactions", "comments"]


@dataclass(frozen=True)
class BenchmarkCase:
    group: str
    platform: str
    operation: Operation
    url: str
    repetition: int
    expected_interactions: int | None = None
    expected_comments: int | None = None
    source_files: tuple[str, ...] = ()

    @property
    def case_id(self) -> str:
        raw = "\0".join((self.group, self.platform, self.operation, self.url, str(self.repetition)))
        return sha256(raw.encode("utf-8")).hexdigest()[:24]


class CountingHttpClient(CurlAsyncHttpClient):
    """Count logical HTTP calls per benchmark task without sharing counters."""

    def __init__(self, **kwargs: Any) -> None:
        super().__init__(**kwargs)
        self._counter: ContextVar[Counter[str] | None] = ContextVar(
            "benchmark_http_counter",
            default=None,
        )

    @asynccontextmanager
    async def count_scope(self) -> AsyncIterator[Counter[str]]:
        counter: Counter[str] = Counter()
        token = self._counter.set(counter)
        try:
            yield counter
        finally:
            self._counter.reset(token)

    async def get(self, url: str, **kwargs: Any) -> Any:
        counter = self._counter.get()
        if counter is not None:
            counter["get"] += 1
        return await super().get(url, **kwargs)

    async def post(self, url: str, **kwargs: Any) -> Any:
        counter = self._counter.get()
        if counter is not None:
            counter["post"] += 1
        return await super().post(url, **kwargs)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="真实 URL 低并发可续跑可靠性测试")
    parser.add_argument("--input-root", type=Path, help="递归查找‘有评论数据*’工作簿")
    parser.add_argument("--special-config", type=Path, help="专项平台 URL JSON")
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument(
        "--mode",
        choices=("dataset", "special", "all"),
        default="all",
    )
    parser.add_argument("--repeats", type=int, default=100, help="专项每个接口的调用次数")
    parser.add_argument("--workers", type=int, default=2)
    parser.add_argument(
        "--platform",
        action="append",
        dest="platforms",
        help="只运行指定平台，可重复传入",
    )
    parser.add_argument(
        "--operation",
        action="append",
        choices=("interactions", "comments"),
        dest="operations",
        help="只运行指定业务接口，可重复传入",
    )
    parser.add_argument(
        "--max-calls-per-second",
        type=float,
        default=2.0,
        help="每个平台的业务接口启动速率；不减少调用次数",
    )
    parser.add_argument("--browser-max-concurrency", type=int, default=1)
    parser.add_argument(
        "--reset-guest-state-on-proxy-change",
        action="store_true",
        help="仅用于可丢弃的快手游客 Profile；代理轮换时重建设备状态",
    )
    parser.add_argument(
        "--disable-browser-fallback",
        action="store_true",
        help="仅测协议阶段；适合已确认浏览器持续验证码后的补充压力测试",
    )
    parser.add_argument(
        "--proxy-mode",
        choices=("direct", "required"),
        default="required",
    )
    return parser


def _integer(value: Any) -> int | None:
    if value in {None, ""}:
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def read_dataset_cases(root: Path) -> list[BenchmarkCase]:
    """Read XLSX content by magic bytes, including workbooks named ``.csv``."""
    merged: dict[str, dict[str, Any]] = {}
    for path in sorted(root.rglob("有评论数据*")):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        try:
            workbook = load_workbook(BytesIO(path.read_bytes()), read_only=True, data_only=True)
        except Exception:
            continue
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows, ())]
        indexes = {name: index for index, name in enumerate(headers)}
        if "URL" not in indexes:
            workbook.close()
            continue
        for row in rows:
            url = str(row[indexes["URL"]] or "").strip()
            if not url.startswith(("http://", "https://")):
                continue
            record = merged.setdefault(
                url,
                {
                    "expected_interactions": None,
                    "expected_comments": None,
                    "source_files": [],
                },
            )
            record["source_files"].append(str(path))
            if "互动量" in indexes:
                record["expected_interactions"] = _integer(row[indexes["互动量"]])
            if "评论数" in indexes:
                record["expected_comments"] = _integer(row[indexes["评论数"]])
        workbook.close()

    cases: list[BenchmarkCase] = []
    for url, record in merged.items():
        try:
            platform, _ = identify_url(url)
        except ValueError:
            platform = "unknown"
        for operation in ("comments", "interactions"):
            cases.append(BenchmarkCase(
                group="dataset",
                platform=platform,
                operation=operation,
                url=url,
                repetition=1,
                expected_interactions=record["expected_interactions"],
                expected_comments=record["expected_comments"],
                source_files=tuple(sorted(set(record["source_files"]))),
            ))
    return cases


def read_special_cases(path: Path, repeats: int) -> list[BenchmarkCase]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    cases: list[BenchmarkCase] = []
    for platform, raw_urls in payload.items():
        urls = [str(url).strip() for url in raw_urls if str(url).strip()]
        if not urls:
            continue
        # Comments run first so a browser-established guest state can be reused
        # by later protocol calls from the same local platform profile.
        for operation in ("comments", "interactions"):
            for index in range(repeats):
                cases.append(BenchmarkCase(
                    group="special",
                    platform=str(platform),
                    operation=operation,
                    url=urls[index % len(urls)],
                    repetition=index + 1,
                ))
    return cases


def _interaction_sum(stats: dict[str, Any]) -> int | None:
    values = [stats.get(name) for name in ("likes", "comments", "shares", "favorites", "reposts")]
    present = [int(value) for value in values if value is not None]
    return sum(present) if present else None


def result_record(
    case: BenchmarkCase,
    result: InteractionResult | CommentPageResult,
    elapsed_seconds: float,
    protocol_calls: Counter[str],
) -> dict[str, Any]:
    dumped = result.model_dump(mode="json")
    record: dict[str, Any] = {
        "case_id": case.case_id,
        "finished_at": datetime.now(timezone.utc).isoformat(),
        "group": case.group,
        "platform": case.platform,
        "operation": case.operation,
        "url": case.url,
        "repetition": case.repetition,
        "elapsed_seconds": round(elapsed_seconds, 6),
        "protocol_get_calls": protocol_calls["get"],
        "protocol_post_calls": protocol_calls["post"],
        "browser_used": str(result.source).startswith("browser"),
        "coverage": result.coverage,
        "source": result.source,
        "reason": result.reason,
        "expected_interactions": case.expected_interactions,
        "expected_comments": case.expected_comments,
        "source_files": list(case.source_files),
        "result": dumped,
    }
    if isinstance(result, InteractionResult):
        stats = result.stats.model_dump()
        record["field_count"] = sum(value is not None for value in stats.values())
        record["interaction_sum"] = _interaction_sum(stats)
        record["outcome"] = "data" if record["field_count"] else "failure"
    else:
        record["returned_comments"] = len(result.comments)
        record["total_comments"] = result.total_comments
        if result.comments:
            record["outcome"] = "data"
        elif result.total_comments == 0 and result.coverage in {"complete", "partial"}:
            record["outcome"] = "valid_empty"
        elif result.total_comments is not None:
            record["outcome"] = "metadata_only"
        else:
            record["outcome"] = "failure"
    return record


def error_record(case: BenchmarkCase, exc: Exception, elapsed_seconds: float) -> dict[str, Any]:
    return {
        "case_id": case.case_id,
        "finished_at": datetime.now(timezone.utc).isoformat(),
        "group": case.group,
        "platform": case.platform,
        "operation": case.operation,
        "url": case.url,
        "repetition": case.repetition,
        "elapsed_seconds": round(elapsed_seconds, 6),
        "outcome": "error",
        "error": f"{type(exc).__name__}: {exc}",
    }


def completed_case_ids(path: Path) -> set[str]:
    completed: set[str] = set()
    if not path.exists():
        return completed
    for line in path.read_text(encoding="utf-8").splitlines():
        try:
            completed.add(str(json.loads(line)["case_id"]))
        except (KeyError, TypeError, ValueError):
            continue
    return completed


async def run(args: argparse.Namespace) -> None:
    if (
        args.workers <= 0
        or args.repeats <= 0
        or args.max_calls_per_second <= 0
        or args.browser_max_concurrency <= 0
    ):
        raise ValueError("workers、repeats、速率和浏览器并发必须大于 0")
    if args.mode in {"dataset", "all"} and args.input_root is None:
        raise ValueError("dataset/all 模式必须提供 --input-root")
    if args.mode in {"special", "all"} and args.special_config is None:
        raise ValueError("special/all 模式必须提供 --special-config")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    results_path = args.output_dir / "results.jsonl"
    cases: list[BenchmarkCase] = []
    if args.mode in {"special", "all"}:
        cases.extend(read_special_cases(args.special_config, args.repeats))
    if args.mode in {"dataset", "all"}:
        cases.extend(read_dataset_cases(args.input_root))
    if args.platforms:
        selected = {platform.strip() for platform in args.platforms}
        cases = [case for case in cases if case.platform in selected]
    if args.operations:
        selected_operations = set(args.operations)
        cases = [case for case in cases if case.operation in selected_operations]
    done = completed_case_ids(results_path)
    pending = [case for case in cases if case.case_id not in done]
    completed_in_scope = sum(case.case_id in done for case in cases)
    print(json.dumps({
        "cases": len(cases),
        "completed": completed_in_scope,
        "pending": len(pending),
    }, ensure_ascii=False))

    settings = Settings()
    provider = None
    if args.proxy_mode == "required":
        if not settings.proxy_51_api_url.strip():
            raise ValueError("required 模式缺少 PROXY_51_API_URL")
        provider = AsyncDailiProxyPool(
            minutes=3,
            pool_size=1,
            max_concurrency_per_proxy=2,
            api_url=settings.proxy_51_api_url,
        )
    session_store = PlatformSessionStore(Path(settings.platform_session_dir))
    client = CountingHttpClient(
        timeout_seconds=settings.request_timeout_seconds,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
            "Accept-Language": "zh-CN,zh;q=0.9",
        },
        proxy_provider=provider,
        proxy_mode=args.proxy_mode,
    )
    browser = None
    if not args.disable_browser_fallback:
        browser = BrowserFallback(
            settings=BrowserFallbackSettings(
                timeout_seconds=settings.browser_timeout_seconds,
                challenge_wait_seconds=settings.browser_challenge_wait_seconds,
                headless=settings.browser_headless,
                max_concurrency=args.browser_max_concurrency,
                profile_dir=Path(settings.browser_profile_dir),
                reset_guest_state_on_proxy_change=args.reset_guest_state_on_proxy_change,
            ),
            proxy_provider=provider,
            session_store=session_store,
            cookies=settings.creator_engagement_cookie.get_secret_value(),
        )
    crawler = EngagementCrawler(
        client=client,
        cookies=settings.creator_engagement_cookie.get_secret_value(),
        proxy_provider=provider,
        proxy_mode=args.proxy_mode,
        browser_fallback=browser,
        session_store=session_store,
    )
    queue: asyncio.Queue[BenchmarkCase] = asyncio.Queue()
    for case in pending:
        queue.put_nowait(case)
    rate_limiters = {
        case.platform: AsyncRequestRateLimiter(args.max_calls_per_second)
        for case in pending
    }
    write_lock = asyncio.Lock()
    completed_now = 0
    started_at = monotonic()

    async def worker() -> None:
        nonlocal completed_now
        while True:
            try:
                case = queue.get_nowait()
            except asyncio.QueueEmpty:
                return
            started = monotonic()
            try:
                await rate_limiters[case.platform].acquire()
                started = monotonic()
                async with client.count_scope() as calls:
                    if case.operation == "comments":
                        result = await crawler.fetch_comments(case.url, case.platform, 1)
                    else:
                        result = await crawler.fetch_interactions(case.url, case.platform)
                record = result_record(case, result, monotonic() - started, calls)
            except Exception as exc:
                record = error_record(case, exc, monotonic() - started)
            async with write_lock:
                with results_path.open("a", encoding="utf-8") as output:
                    output.write(json.dumps(record, ensure_ascii=False) + "\n")
                    output.flush()
                completed_now += 1
                if completed_now % 10 == 0 or completed_now == len(pending):
                    print(json.dumps({
                        "finished": completed_now,
                        "pending": len(pending) - completed_now,
                        "last": {"platform": case.platform, "operation": case.operation, "outcome": record["outcome"]},
                        "elapsed_seconds": round(monotonic() - started_at, 1),
                    }, ensure_ascii=False), flush=True)
            queue.task_done()

    try:
        await asyncio.gather(*(worker() for _ in range(min(args.workers, max(1, len(pending))))))
    finally:
        await client.aclose()
        run_record = {
            "finished_at": datetime.now(timezone.utc).isoformat(),
            "mode": args.mode,
            "proxy_mode": args.proxy_mode,
            "max_calls_per_second": args.max_calls_per_second,
            "browser_max_concurrency": args.browser_max_concurrency,
            "reset_guest_state_on_proxy_change": args.reset_guest_state_on_proxy_change,
            "browser_fallback_enabled": not args.disable_browser_fallback,
            "duration_seconds": round(monotonic() - started_at, 6),
            "completed_calls": completed_now,
            "stats": asdict(provider.stats) if provider is not None else None,
        }
        with (args.output_dir / "benchmark_runs.jsonl").open("a", encoding="utf-8") as output:
            output.write(json.dumps(run_record, ensure_ascii=False) + "\n")
        if provider is not None:
            await provider.close()


def main() -> None:
    args = build_parser().parse_args()
    configure_logging(Settings())
    asyncio.run(run(args))


if __name__ == "__main__":
    main()
